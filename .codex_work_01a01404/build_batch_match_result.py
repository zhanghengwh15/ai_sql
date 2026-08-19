import csv
import json
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

CSV_PATH = Path('/Users/zhangheng/database/mysql/poit_data_center_dater_inbound_finsh_data.csv')
XLSX_PATH = Path('/Users/zhangheng/Downloads/工单投料数据260819101057.xlsx')
OUTPUT_PATH = Path('/Users/zhangheng/database/ai_sql/outputs/01a01404-c93d-7231-8b66-0cd0a00c21f9/历史错误数据.xlsx')

NAVY = '1F4E78'
BLUE = 'D9EAF7'
LIGHT_BLUE = 'EAF3F8'
RED = 'FCE4D6'
GRAY = 'F2F2F2'
BORDER = 'B7C9D6'
RED_TEXT = '9C0006'


def parse_all_details():
    details = []
    with CSV_PATH.open('r', encoding='utf-8-sig', newline='') as source:
        for csv_row, row in enumerate(csv.reader(source), start=1):
            document_no = row[0]
            for detail_index, item in enumerate(json.loads(row[1]), start=1):
                details.append({
                    'CSV行号': csv_row,
                    '入库单号': document_no,
                    'JSON明细序号': detail_index,
                    '入库物料编码': str(item.get('materialCode', '')).strip(),
                    '入库批次号': str(item.get('batchNo', '')).strip(),
                    '入库数量': item.get('quantity', ''),
                    '入库时间': item.get('entryTime', ''),
                    '入库库位编码': str(item.get('entryStoreLocationCode', '')).strip(),
                })
    return details


def add_title(sheet, title, end_column):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, title)
    cell.fill = PatternFill('solid', fgColor=NAVY)
    cell.font = Font(color='FFFFFF', bold=True, size=16)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    sheet.row_dimensions[1].height = 28


def style_header(sheet, row, column_count):
    thin = Side(style='thin', color=BORDER)
    for column in range(1, column_count + 1):
        cell = sheet.cell(row, column)
        cell.fill = PatternFill('solid', fgColor=BLUE)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet.row_dimensions[row].height = 28


def write_table_sheet(workbook, name, title, headers, records, widths, text_headers=(), number_headers=()):
    sheet = workbook.create_sheet(name)
    add_title(sheet, title, len(headers))
    for column, header in enumerate(headers, start=1):
        sheet.cell(2, column, header)
    style_header(sheet, 2, len(headers))

    for row_index, record in enumerate(records, start=3):
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row_index, column, record.get(header, ''))
            cell.alignment = Alignment(vertical='center')
            if header in text_headers:
                cell.number_format = '@'
            elif header in number_headers:
                cell.number_format = '#,##0.0000'

    if records:
        ref = f'A2:{get_column_letter(len(headers))}{len(records) + 2}'
        table = Table(displayName=f'{name}Table', ref=ref)
        table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)

    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = 'A3'
    sheet.sheet_view.showGridLines = False
    return sheet


details = parse_all_details()
materials = pd.read_excel(XLSX_PATH, sheet_name='工单投料数据', dtype=str, keep_default_na=False)
materials['物料批次号'] = materials['物料批次号'].astype(str).str.strip()
material_records = materials.to_dict(orient='records')

target_by_batch = {}
for record in material_records:
    target_by_batch.setdefault(record['物料批次号'], []).append(record)

matched_details = [detail for detail in details if detail['入库批次号'] in target_by_batch]
unmatched_details = [detail for detail in details if detail['入库批次号'] not in target_by_batch]
matched_rows = []
for detail in matched_details:
    for target in target_by_batch[detail['入库批次号']]:
        matched_rows.append({**detail, **target})

source_batches = Counter(detail['入库批次号'] for detail in details)
matched_batches = sorted(batch for batch in source_batches if batch in target_by_batch)

workbook = Workbook()
workbook.remove(workbook.active)

summary = workbook.create_sheet('对比汇总')
add_title(summary, '按批次号匹配分析', 6)
summary_rows = [
    ('CSV 数据行数', 5),
    ('全部 JSON 明细数', len(details)),
    ('JSON 唯一批次号数', len(source_batches)),
    ('完全匹配批次号数', len(matched_batches)),
    ('命中 JSON 明细数', len(matched_details)),
    ('命中工单投料行数', sum(len(target_by_batch[batch]) for batch in matched_batches)),
    ('批次号关联展开行数', len(matched_rows)),
    ('未匹配 JSON 明细数', len(unmatched_details)),
    ('匹配规则', '仅批次号完全一致：JSON batchNo = 工单投料数据 物料批次号'),
]
thin = Side(style='thin', color=BORDER)
for row, (label, value) in enumerate(summary_rows, start=3):
    left = summary.cell(row, 1, label)
    left.fill = PatternFill('solid', fgColor=BLUE)
    left.font = Font(bold=True)
    left.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    right = summary.cell(row, 2, value)
    right.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    right.alignment = Alignment(vertical='center', wrap_text=True)
    if label in {'完全匹配批次号数', '命中 JSON 明细数', '命中工单投料行数', '批次号关联展开行数'}:
        right.fill = PatternFill('solid', fgColor=LIGHT_BLUE)
summary.cell(13, 1, '来源文件')
summary.cell(13, 1).fill = PatternFill('solid', fgColor=LIGHT_BLUE)
summary.cell(13, 1).font = Font(bold=True)
summary.cell(14, 1, 'CSV')
summary.cell(14, 2, CSV_PATH.name)
summary.cell(15, 1, '工单投料数据')
summary.cell(15, 2, XLSX_PATH.name)
for row in (13, 14, 15):
    for column in (1, 2):
        summary.cell(row, column).border = Border(top=thin, bottom=thin, left=thin, right=thin)
summary.column_dimensions['A'].width = 25
summary.column_dimensions['B'].width = 76
summary.sheet_view.showGridLines = False

detail_headers = ['CSV行号', '入库单号', 'JSON明细序号', '入库物料编码', '入库批次号', '入库数量', '入库时间', '入库库位编码']
write_table_sheet(
    workbook, '全部JSON明细', 'CSV 第二列 JSON 全部展开明细', detail_headers, details,
    [11, 20, 13, 16, 24, 14, 25, 15],
    text_headers={'入库单号', '入库物料编码', '入库批次号', '入库库位编码'}, number_headers={'入库数量'},
)

matched_headers = detail_headers + [
    '生产工单号', '工单模板', '投料工厂单元', '投料日期', '投料班次', '投料物料编码',
    '投料物料名称', '投料计量单位', '投料数量', '退料数量', '投料状态', '退料备注',
    '单据编码', '投料时间', '投料人', '投料班组', '物料批次号', '出库仓库/库位',
    '容器', '产品物料名称', '计划产量', '产品计量单位', '产品规格描述', 'BOM版本', '工序工艺', '供应商',
]
write_table_sheet(
    workbook, '匹配工单明细', '仅按批次号匹配的工单投料行', matched_headers, matched_rows,
    [11, 20, 13, 16, 24, 14, 25, 15, 18, 16, 18, 13, 15, 16, 16, 24, 12, 14, 12, 15, 22, 20, 22, 22, 16, 18, 15, 22, 14, 25, 14, 16, 20, 12],
    text_headers={'入库单号', '入库物料编码', '入库批次号', '入库库位编码', '生产工单号', '投料物料编码', '单据编码', '物料批次号'}, number_headers={'入库数量', '投料数量', '退料数量', '计划产量'},
)

write_table_sheet(
    workbook, '未匹配JSON明细', '未匹配的 JSON 明细', detail_headers, unmatched_details,
    [11, 20, 13, 16, 24, 14, 25, 15],
    text_headers={'入库单号', '入库物料编码', '入库批次号', '入库库位编码'}, number_headers={'入库数量'},
)

batch_headers = ['匹配批次号', 'JSON明细次数', '工单投料行数', '关联展开行数']
batch_records = [
    {
        '匹配批次号': batch,
        'JSON明细次数': source_batches[batch],
        '工单投料行数': len(target_by_batch[batch]),
        '关联展开行数': source_batches[batch] * len(target_by_batch[batch]),
    }
    for batch in matched_batches
]
write_table_sheet(
    workbook, '匹配批次号汇总', '匹配批次号汇总', batch_headers, batch_records,
    [26, 16, 16, 18], text_headers={'匹配批次号'},
)

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
workbook.save(OUTPUT_PATH)

check = load_workbook(OUTPUT_PATH, read_only=True, data_only=False)
assert check['全部JSON明细'].max_row == len(details) + 2
assert check['匹配工单明细'].max_row == len(matched_rows) + 2
assert check['未匹配JSON明细'].max_row == len(unmatched_details) + 2
assert check['匹配批次号汇总'].max_row == len(batch_records) + 2
print({
    'output': str(OUTPUT_PATH),
    'json_detail_count': len(details),
    'matched_batch_count': len(matched_batches),
    'matched_json_detail_count': len(matched_details),
    'matched_row_count': len(matched_rows),
    'unmatched_detail_count': len(unmatched_details),
})
